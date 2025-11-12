import sys
import openpyxl as ox
import pandas as pd
from openpyxl.styles import Font, Border, Side

def trim_dict(d , n):
    return dict(list(d.items())[:n])

def prosecc_dataset(filepath):
    real_df = pd.read_csv(filepath, names=['name', 'sal_f', 'sal_to','sal_c', 'city', 'date']).dropna()
    df = real_df.assign(date = lambda x: x['date'].str[:4])

    df['mid_sal'] = (df[['sal_f', 'sal_to']]
                     .mean(axis=1)
                     .dropna()
                     .round(0).astype(int)
                     )
    
    df = df.drop(columns='sal_f').drop(columns='sal_to')
    s_mid_sal= (df
                .groupby('date')['mid_sal']
                .mean()
                .dropna()
                .round(0).astype(int)
                )
    df = df[df['sal_c'].str.contains('RUR', na=True)]
    city_mid_sal = (df
        .groupby('city')['mid_sal']
        .mean()
        .dropna()
        .round(0).astype(int)
    ).sort_values(ascending=False)
    df = df[['name', 'date']].dropna()

    city_vac = pd.Series(index=city_mid_sal.index)
    for city, _ in city_vac.items():
        count = (real_df['city'] == city).sum()
        city_vac[city] = count

    s_vac = pd.Series(index=s_mid_sal.index)
    for date, _ in s_mid_sal.items():
        count = (df['date'] == date).sum()
        s_vac[date] = count

    s_vac = s_vac.dropna()
    all_vac_count = s_vac.sum()
    s_mid_sal = s_mid_sal.to_dict()
    s_vac = s_vac.to_dict() 
    
    city_vac_p = city_vac.apply(lambda x: round((x / all_vac_count)*100, 2)).sort_values(ascending=False)
    
    city_vac_p = city_vac_p.to_dict()
    d = {k: v for k, v in city_vac_p.items() if v > 1}
    city_mid_sal = city_mid_sal.to_dict()
    city_mid_sal = {k: city_mid_sal[k] for k in city_mid_sal if k in d}
    print(trim_dict(city_vac_p, 10))
    print(trim_dict(city_mid_sal, 10))
    return {year: (s_mid_sal[year], s_vac[year]) for year, _ in s_vac.items()}, city_vac_p, city_mid_sal
def create_xls(comb):
    
    
    years_data, city_share, city_salary = comb
    wb = ox.Workbook()
    ws = wb.active

    year_stat = wb.create_sheet(title='Статистика по годам', index=0)
    city_stat = wb.create_sheet(title='Статистика по городам', index=1)
    del wb['Sheet']    
    # year_stat
    year_stat.append(['Год', 'Средняя зарплата','Количество вакансий'])
    #font
    topic_font = Font(name='Calibri', size=11, bold=True)
    for i in ['A', 'B', 'C']:
        year_stat[f'{i}1'].font = topic_font
    # filling
    for key, ar in trim_dict(years_data, 18).items():
        year_stat.append([key, ar[0], ar[1]])
    # border
    thin_side = Side(style="thin", color="000000")
    border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    for i in range(1, 18):
        year_stat[f'A{i}'].border = border
        year_stat[f'B{i}'].border = border
        year_stat[f'C{i}'].border = border

    year_stat.column_dimensions[f'A'].width = 6.29
    year_stat.column_dimensions[f'B'].width = 20.29
    year_stat.column_dimensions[f'C'].width = 20.29
    # city_stat
    city_stat.column_dimensions[f'A'].width = 20.29
    city_stat.column_dimensions[f'B'].width = 20.29
    city_stat.column_dimensions[f'C'].width = 2.29
    city_stat.column_dimensions[f'D'].width = 20.29
    city_stat.column_dimensions[f'E'].width = 20.29
    city_stat.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий, %'])

    for i in ['A', 'B', 'D', 'E']:
       city_stat[f'{i}1'].font = topic_font
    d1 = {}

    d1 = dict(sorted(city_salary.items(), key=lambda x: x[1], reverse=True))
    i = 1
    for (k1, v1), (k2, v2) in zip(d1.items(), city_share.items()):
        city_stat.append([k1, v1,'' , k2, v2])
        if i == 10:
             break
        i += 1

    for i in range(1, 12):
        city_stat[f'A{i}'].border = border
        city_stat[f'B{i}'].border = border
        city_stat[f'D{i}'].border = border
        city_stat[f'E{i}'].border = border

    wb.save('student_works/report.xlsx')
if __name__ == '__main__':
    try:
        if sys.argv[1] == 'test':
            create_xls(prosecc_dataset(r'test/vacancies.csv'))
    except IndexError:
        create_xls(prosecc_dataset(r'vacancies.csv'))
        
